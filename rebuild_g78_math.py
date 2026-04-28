#!/usr/bin/env python3
"""Comprehensive rebuild of G7 and G8 math content from Excel.

Strategy:
  1. Delete every existing math block whose grades are EXACTLY {G7}, {G8},
     or {G7,G8} (these are the misaligned blocks).
  2. For each Excel math slot, gather (title, std) per grade and group:
       - If G7 and G8 share content (and it differs from G5/G6) -> one
         G7+G8 block.
       - If G7 ≠ G8 -> two blocks (G7-only and G8-only).
       - If all four match -> handled by existing G5+G6+G7+G8 block,
         skip (we don't touch shared 4-grade blocks).
       - Otherwise tag with the grade subset that has that exact content.
  3. Build PREREQS chains for the new G7-only and G8-only math tags
     (chronological order = chain), each rooted at math-diagnostic.
  4. Handle Sep 15 cog/challenge swap as a special case (G8 has those at
     swapped times).
  5. Update state.json + PUBLISHED_STATE in index.html. Bump
     __schema_version so the version-gate forces it into Supabase.
"""

import openpyxl, datetime, json, re, sys

REPO = '/Users/ktusiime/Desktop/DLA/Forge/Code/CohortCalendar'
EXCEL = '/Users/ktusiime/Downloads/Lunch Challenge Cohort Calendar latest version.xlsx'

WEEK_STARTS = {
    1: '2026-09-07', 2: '2026-09-14', 3: '2026-09-21', 4: '2026-09-28',
    5: '2026-10-05', 6: '2026-10-12', 7: '2026-10-19', 8: '2026-10-26',
}
SUBJ_TO_TP = {
    'math': 'math', 'biology': 'biology', 'chem': 'chem',
    'lab': 'chem-lab', 'ela': 'ela', 'challenge': 'challenge',
    'cog': 'cog-check', 'lunch': 'lunch', 'movement': 'movement',
    'diag': 'diagnostic',
}

def parse_date(s, year=2026):
    return datetime.datetime.strptime(f'{s} {year}', '%b %d %Y').date()

def date_to_wd(d):
    for w, start in WEEK_STARTS.items():
        ws = datetime.date.fromisoformat(start)
        if ws <= d <= ws + datetime.timedelta(days=4):
            return (w, (d - ws).days)
    return None

def time_to_min(t):
    h, m = t.split(':'); h, m = int(h), int(m)
    if h < 9: h += 12
    return (h - 9) * 60 + m

def parse_time_range(t):
    s, e = t.split('-')
    return time_to_min(s), time_to_min(e)

def gather_grade(wb, sheet):
    ws = wb[sheet]
    out = []
    for r in range(5, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, 8)]
        sess = row[3]
        if not sess or sess == 'DAY TOTAL': continue
        if not row[0] or not row[2]: continue
        if isinstance(row[0], str) and row[0].startswith('Week'): continue
        try:
            d = parse_date(row[0])
            wd = date_to_wd(d)
            if not wd: continue
            s_min, e_min = parse_time_range(row[2])
        except Exception:
            continue
        out.append({
            'w': wd[0], 'd': wd[1], 's': s_min, 'dur': e_min - s_min,
            'ttl': sess.strip(),
            'subj': row[4],
            'std': [x.strip() for x in (row[5] or '').split(',') if x.strip()],
        })
    return out

def title_to_tag(ttl, suffix):
    slug = re.sub(r'[^a-z0-9]+', '-', ttl.lower()).strip('-')
    if len(slug) > 32: slug = slug[:32].rstrip('-')
    return f'{slug}-{suffix}'

def make_block(s, grades, bid):
    return {
        'id': bid,
        'w': s['w'], 'd': s['d'], 's': s['s'], 'dur': s['dur'],
        'ttl': s['ttl'],
        'std': s['std'],
        'desc': '',
        'std_defensible': [],
        'tp': SUBJ_TO_TP.get(s['subj'], 'challenge'),
        'tag': None,  # set below
        'anc': False,
        'pin': False,
        'grades': grades,
        'locked': False,
    }

def main():
    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    grades_data = {g: gather_grade(wb, f'Grade {g[1]} Schedule')
                   for g in ['G5', 'G6', 'G7', 'G8']}

    state = json.load(open(f'{REPO}/state.json'))
    blocks = state['blocks']

    # Phase 1: drop misaligned math blocks (those exactly tagged {G7}, {G8},
    # or {G7,G8}). 4-grade and G5/G6 math blocks are preserved.
    drop = lambda b: (
        b.get('tp') == 'math'
        and sorted(b.get('grades', [])) in ([['G7']], [['G8']], [['G7','G8']])[0:3]
    )
    # The lambda above is wrong syntactically — rewrite:
    def drop_misaligned(b):
        if b.get('tp') != 'math': return False
        g = sorted(b.get('grades', []))
        return g == ['G7'] or g == ['G8'] or g == ['G7', 'G8']
    n_before = len(blocks)
    dropped = [b['id'] for b in blocks if drop_misaligned(b)]
    blocks = [b for b in blocks if not drop_misaligned(b)]
    print(f'Phase 1: dropped {len(dropped)} misaligned math blocks: {dropped[:6]}{"..." if len(dropped)>6 else ""}')

    # Phase 2: per-slot rebuild of G7/G8 (and G7+G8 shared) math.
    # Build slot -> grade -> session mapping.
    slot_to_grade_session = {}  # (w,d,s,dur) -> {'G5': sess, ...}
    for grade, sessions in grades_data.items():
        for s in sessions:
            if s['subj'] != 'math': continue
            key = (s['w'], s['d'], s['s'], s['dur'])
            slot_to_grade_session.setdefault(key, {})[grade] = s

    # For each math slot, group grades by content.
    next_id_n = 0
    used_ids = {b['id'] for b in blocks}
    def new_id(pre):
        nonlocal next_id_n
        while True:
            next_id_n += 1
            cand = f'{pre}_{next_id_n:03d}'
            if cand not in used_ids:
                used_ids.add(cand)
                return cand

    g7_chain = []   # [(w,d,s,tag), ...] for chronological prereq chain
    g8_chain = []
    added = []

    for key, gmap in slot_to_grade_session.items():
        w, d, s, dur = key
        # We only act if G7 or G8 is in this slot's Excel data.
        if 'G7' not in gmap and 'G8' not in gmap:
            continue

        # Group by (ttl, tuple(std))
        groups = {}
        for g, sess in gmap.items():
            ck = (sess['ttl'], tuple(sess['std']))
            groups.setdefault(ck, []).append(g)

        # Find what's already at this slot (math blocks remaining after Phase 1).
        existing = [b for b in blocks if b['w']==w and b['d']==d and b['s']==s and b['tp']=='math']

        for (ttl, std_tup), grade_list in groups.items():
            grade_set = sorted(grade_list)
            # If this group includes G5 or G6, it's likely covered by an
            # existing G5/G6/G5+G6 block — skip (don't touch G5/G6 path).
            if 'G5' in grade_set or 'G6' in grade_set:
                continue

            # Now grade_set is a subset of {G7, G8}. Create one block.
            sess_template = gmap[grade_list[0]]
            # Pick suffix for tag based on which grades.
            if grade_set == ['G7', 'G8']:
                suffix = '78'
            elif grade_set == ['G7']:
                suffix = '7'
            elif grade_set == ['G8']:
                suffix = '8'
            else:
                continue

            nb = make_block(sess_template, grade_set, new_id('blk_n'))
            nb['tag'] = title_to_tag(ttl, suffix)
            added.append(nb)

            if suffix == '7':
                g7_chain.append((w, d, s, nb['tag']))
            elif suffix == '8':
                g8_chain.append((w, d, s, nb['tag']))
            # G7+G8 shared math doesn't need a chain entry (those tags exist
            # but their prereqs come from the canonical PREREQS map).

    blocks.extend(added)
    print(f'Phase 2: added {len(added)} fresh G7/G8 math blocks (G7-only, G8-only, or G7+G8 shared)')

    # Phase 3: Sep 15 cog/challenge swap. Per Excel:
    #   G7 (and G5, G6): cog setup at 10:10 (s=70, dur=45), challenge at 11:00 (s=120, dur=35)
    #   G8: challenge at 10:10, cog setup at 11:00
    # Strategy: strip G8 from existing blocks at those slots, create G8-only
    # blocks at the swapped times.
    SEP15 = 2  # week
    SWAP_PAIRS = [
        # (existing slot, existing title fragment, new G8 content from Excel)
        ((SEP15, 1, 70),  'Cognitive baseline'),
        ((SEP15, 1, 120), 'FoodData Central'),
    ]
    # Find G8 sessions for those swapped times: gmap from Excel.
    # G8 at (w=2, d=1, s=70) is challenge "FoodData Central: interpretation"
    # G8 at (w=2, d=1, s=120) is cog "Cognitive baseline: setup"
    g8_at_70 = next((s for s in grades_data['G8']
                     if s['w']==2 and s['d']==1 and s['s']==70), None)
    g8_at_120 = next((s for s in grades_data['G8']
                      if s['w']==2 and s['d']==1 and s['s']==120), None)
    swap_actions = 0
    for (w, d, s), _ in SWAP_PAIRS:
        for b in blocks:
            if b['w']==w and b['d']==d and b['s']==s and 'G8' in b.get('grades', []):
                b['grades'] = [g for g in b['grades'] if g != 'G8']
                swap_actions += 1
    if g8_at_70:
        nb = make_block(g8_at_70, ['G8'], new_id('blk_g8'))
        nb['tag'] = title_to_tag(g8_at_70['ttl'], '8')
        blocks.append(nb)
        swap_actions += 1
    if g8_at_120:
        nb = make_block(g8_at_120, ['G8'], new_id('blk_g8'))
        nb['tag'] = title_to_tag(g8_at_120['ttl'], '8')
        blocks.append(nb)
        swap_actions += 1
    print(f'Phase 3: Sep 15 cog/challenge swap — {swap_actions} surgical actions')

    # Phase 4: build PREREQS chains
    g7_chain.sort()
    g8_chain.sort()
    new_prereqs = {}
    for i, (_, _, _, tag) in enumerate(g7_chain):
        new_prereqs[tag] = ['math-diagnostic'] if i == 0 else [g7_chain[i-1][3]]
    for i, (_, _, _, tag) in enumerate(g8_chain):
        new_prereqs[tag] = ['math-diagnostic'] if i == 0 else [g8_chain[i-1][3]]
    print(f'Phase 4: G7 chain={len(g7_chain)}, G8 chain={len(g8_chain)}')

    # Phase 5: write state.json + index.html
    state['blocks'] = blocks
    state['__schema_version'] = datetime.datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ')
    json.dump(state, open(f'{REPO}/state.json', 'w'), indent=2)
    print(f'Phase 5: state.json written. blocks={len(blocks)} (was {n_before})')
    print(f'         __schema_version = {state["__schema_version"]}')

    # Inject into index.html using string replacement (not regex sub) to avoid
    # backslash-escape issues from JSON content.
    index_path = f'{REPO}/index.html'
    src = open(index_path).read()
    new_state_json = json.dumps(state, separators=(', ', ': '))

    # Find and replace PUBLISHED_STATE block manually
    start = src.find('const PUBLISHED_STATE = ')
    if start < 0:
        print('ERROR: PUBLISHED_STATE not found'); sys.exit(1)
    # Find end: matching '};' after start
    end = src.find('};', start) + len('};')
    src = src[:start] + f'const PUBLISHED_STATE = {new_state_json};' + src[end:]

    # Update PREREQS: parse, merge new entries, write back
    pre_match = re.search(r'const PREREQS = (\{.*?\});', src, re.DOTALL)
    pre = json.loads(pre_match.group(1))
    pre.update(new_prereqs)
    pre_compact = json.dumps(pre, separators=(', ', ': '))
    pre_start = pre_match.start()
    pre_end = pre_match.end()
    src = src[:pre_start] + f'const PREREQS = {pre_compact};' + src[pre_end:]

    open(index_path, 'w').write(src)
    print(f'Phase 5: index.html updated.')

if __name__ == '__main__':
    main()
